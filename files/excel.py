import openpyxl
import os


my_values = [
    ['A1', 'A2', 'A3', 'A4', 'A5'],
    ['B1', 'B2', 'B3', 'B4', 'B5'],
    ['C1', 'C2', 'C3', 'C4', 'C5']
]


def create_workbook(name, values):
    """
    :param name: the name with which the file will be saved
    :param values:
    the values multidimensional-array must look like this
    values = [
        [ 'A1', 'A2', 'A3', 'A4', 'A5', ... ],
        [ 'B1', 'B2', 'B3', 'B4', 'B5', ... ],
        [ 'C1', 'C2', 'C3', 'C4', 'C5', ... ]
    ]
    :return: True is successfully saved
    """
    wb = openpyxl.Workbook()
    # for v_row in values:
    #     for v_col in v_row:
    #         print('row: {}\tcol = {}'.format(v_row, v_col), end=' ')
    #     print()
    #
    # print('|Numeric single range values')
    # for i in values:
    #     for j in range(len(i)):
    #         print('val[{}] = {}'.format(j, i[j]))
    #     print()
    # print('|Numeric range values')

    sheet = wb['Sheet']
    for n_row in range(len(values)):
        print('val[{}] = {}'.format(n_row, values[n_row]))
        for n_col in range(len(values[n_row])):
            sheet.cell(row=n_col+1, column=n_row+1).value = values[n_row][n_col]
            print('val[{}][{}] = {}'.format(n_row, n_col, sheet.cell(row=n_row+1, column=n_col+1).value))

    os.chdir('/home/pit/Documents/python_projects/python-structure/files/')
    wb.save(filename=name)


def create_new_sheet(file_path, new_sheet_name, new_s_index=None):
    wb = get_workbook(file_path)
    new_sheet = wb.create_sheet(new_sheet_name, new_s_index)
    return new_sheet.title


def get_workbook(file_path):
    return openpyxl.load_workbook(file_path)


def get_sheet(file_path, s_name):
    workbook = get_workbook(file_path)
    # sheet = workbook.get_sheet_by_name(s_name) [deprecated]
    return workbook[s_name]


def get_sheets(file_path):
    workbook = get_workbook(file_path)
    # sheets = workbook.get_sheet_names() [deprecated]
    return workbook.sheetnames


def get_cell(file_path, s_name, cell_name):
    sheet = get_sheet(file_path, s_name)

    # for cell_name = 'A1' --> cell = sheet.cell(row=1, column=1)
    cell = sheet[cell_name]
    # write a def that check for the type and return it in a more appropriate way
    return str(cell.value)


def get_cells(file_path, s_name, row, column):
    sheet = get_sheet(file_path, s_name)
    for i in range(1, column + 1):
        for j in range(1, row + 1):
            msg = '{}-{}:\t{}'.format(i, j, sheet.cell(row=j, column=i).value)
            if j == 1:
                print(msg.upper())
            else:
                print(msg)
        print('')


if __name__ == '__main__':
    ex_file = 'excel_file.xlsx'
    print(get_sheet(ex_file, 'Sheet1'))
    print(get_sheets(ex_file))
    print(get_cell(ex_file, 'Sheet1', 'A1'))
    get_cells(ex_file, 'Sheet1', 4, 5)
    create_workbook('test.xlsx', my_values)
    exit(0)
